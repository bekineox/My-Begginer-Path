import os
import datetime
import glob
import threading
import time

from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

import schedule

# === Load your BOT token from environment ===
BOT_TOKEN = "8190151520:AAF1-fiq4qxJHXEfmnqJRXnNepMYHa2FwDY"

# === Admin user ID (replace this with your own ID) ===
ADMIN_ID = 6912441144  # ← Replace this with your Telegram user ID (a number)

# === In-memory tracking ===
attendance_data = {}
awaiting_users = set()

# === /start ===
from telegram import ReplyKeyboardMarkup

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["📝 Mark"],
        ["📋 My Attendance"],
        ["📊 Report"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "👋 Welcome to the Attendance Bot!\n"
        "Choose an option below:",
        reply_markup=reply_markup
    )
# === /mark ===
async def mark(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in awaiting_users:
        await update.message.reply_text("⚠️ You're already being asked. Please reply with:\nFull Name, Student ID")
    else:
        awaiting_users.add(user_id)
        await update.message.reply_text("📝 Please enter your full name and student ID:\nFull Name, Student ID")

# === Handle name and ID input ===
async def collect_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from datetime import datetime as dt

    user_id = update.effective_user.id
    text = update.message.text.strip()
    today = datetime.date.today().isoformat()

    if user_id not in awaiting_users:
        return

    if ',' not in text:
        await update.message.reply_text("❌ Format error. Use:\nFull Name, Student ID", parse_mode="Markdown")
        return

    name, student_id = map(str.strip, text.split(",", 1))
    if not name or not student_id:
        await update.message.reply_text("❌ Missing name or ID.")
        return

    now = dt.now()
    time_str = now.strftime("%Y-%m-%d %H:%M:%S")
    record = {"user_id": user_id, "name": name, "student_id": student_id, "time": time_str}
    attendance_data.setdefault(today, []).append(record)
    awaiting_users.remove(user_id)

    context.user_data["last_name_id"] = f"{name}, {student_id}"

    os.makedirs("attendance_logs", exist_ok=True)

    # Save to TXT
    txt_file = f"attendance_logs/attendance_{today}.txt"
    with open(txt_file, "a") as f:
        f.write(f"{time_str} - {name}, {student_id}\n")

    # Save to Excel
    excel_file = f"attendance_logs/attendance_{today}.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Date", "Time", "Full Name", "Student ID"])
    else:
        wb = load_workbook(excel_file)
        ws = wb.active

    ws.append([today, time_str, name, student_id])
    wb.save(excel_file)

    await update.message.reply_text("✅ Attendance recorded. Thank you!")

# === /myattendance ===
async def myattendance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name_id = context.user_data.get("last_name_id")
    if not name_id:
        await update.message.reply_text("❌ You haven't marked attendance yet.")
        return

    records = []
    for file in glob.glob("attendance_logs/attendance_*.xlsx"):
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, time_, name, sid = row
            if f"{name}, {sid}" == name_id:
                records.append((date, time_))

    if not records:
        await update.message.reply_text("🔍 No records found.")
    else:
        msg = f"📋 Your Attendance Records ({len(records)}):\n\n"
        msg += "\n".join([f"✅ {d} at {t}" for d, t in records])
        await update.message.reply_text(msg)
        # === /report (admin only) ===
async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔ You're not allowed to use this command.")
        return

    today = datetime.date.today().isoformat()
    file_path = f"attendance_logs/attendance_{today}.xlsx"

    if not os.path.exists(file_path):
        await update.message.reply_text("📭 No attendance data for today.")
        return

    wb = load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        await update.message.reply_text("📭 No entries found.")
        return

    msg = f"📋 *Today's Attendance Report* ({len(rows)} students):\n\n"
    for i, (date, time_, name, sid) in enumerate(rows, 1):
        msg += f"{i}. {name} ({sid}) at {time_}\n"

    for part in [msg[i:i+4000] for i in range(0, len(msg), 4000)]:
        await update.message.reply_text(part, parse_mode="Markdown")

# === Background scheduler (placeholder) ===
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(60)

# === MAIN ===
if __name__ == '__main__':
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Command handlers (for commands like /start, /mark, etc.)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("mark", mark))
    app.add_handler(CommandHandler("myattendance", myattendance))
    app.add_handler(CommandHandler("report", report))

    # Button text handlers (when users tap on buttons instead of typing commands)
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex("📝 Mark"), mark))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex("📋 My Attendance"), myattendance))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex("📊 Report"), report))

    # Handles text replies with name and ID
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, collect_info))

    # Start background scheduler (you can leave this as is)
    threading.Thread(target=run_scheduler, daemon=True).start()

    print("✅ Bot is running...")
    app.run_polling()