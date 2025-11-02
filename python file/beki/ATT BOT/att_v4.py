import os
import datetime
import sqlite3
import logging
import threading
import time
from typing import Optional, List

from openpyxl import Workbook, load_workbook
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)

import schedule

# === Global bot status ===
bot_active = False  # Tracks if attendance bot is active

# === Configuration ===
BOT_TOKEN = "8469768440:AAFqN4d9cureq3g0f7N1-GGBM0CvZ69fK-U"
ADMIN_ID = 8191641097  # Replace with your Telegram user ID

# === Database Setup ===
DB_PATH = "attendance.db"

def init_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            full_name TEXT NOT NULL,
            student_id TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            student_id TEXT,
            registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# === Conversation States ===
NAME, STUDENT_ID, DELETE_STUDENT_ID = range(3)

# === Logging Setup ===
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# === Attendance Bot Class ===
class AttendanceBot:
    def __init__(self):
        self.awaiting_registration = set()

    # --- Helper methods ---
    def get_user_info(self, user_id: int) -> Optional[tuple]:
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
            user = cursor.fetchone()
            conn.close()
            return user
        except sqlite3.Error as e:
            logger.error(f"Database error in get_user_info: {e}")
            return None

    def student_id_exists(self, student_id: str) -> bool:
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM users WHERE student_id = ?", (student_id,))
            exists = cursor.fetchone() is not None
            conn.close()
            return exists
        except sqlite3.Error as e:
            logger.error(f"Database error in student_id_exists: {e}")
            return False

    def already_marked_today(self, user_id: int, date: str) -> bool:
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT 1 FROM attendance WHERE user_id = ? AND date = ?",
                (user_id, date)
            )
            marked = cursor.fetchone() is not None
            conn.close()
            return marked
        except sqlite3.Error as e:
            logger.error(f"Database error in already_marked_today: {e}")
            return False

    def save_to_excel(self, date: str, time: str, name: str, student_id: str):
        try:
            os.makedirs("attendance_logs", exist_ok=True)
            excel_file = f"attendance_logs/attendance_{date}.xlsx"

            if not os.path.isfile(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws.append(["Date", "Time", "Full Name", "Student ID"])
            else:
                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Attendance"
                    ws.append(["Date", "Time", "Full Name", "Student ID"])

            ws.append([date, time, name, student_id])
            temp_file = f"{excel_file}.tmp"
            wb.save(temp_file)
            wb.close()
            os.replace(temp_file, excel_file)

        except Exception as e:
            logger.error(f"Excel save error: {e}")

    def generate_detailed_report(self, date: str, records: List[tuple]):
        try:
            os.makedirs("attendance_logs", exist_ok=True)
            excel_file = f"attendance_logs/attendance_{date}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            ws.append(["No.", "Full Name", "Student ID", "Time", "Date"])
            for i, (name, student_id, time) in enumerate(records, 1):
                ws.append([i, name, student_id, time, date])
            
            wb.save(excel_file)
        except Exception as e:
            logger.error(f"Excel report error: {e}")

    def get_all_users(self) -> List[int]:
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT user_id FROM users")
            users = [row[0] for row in cursor.fetchall()]
            conn.close()
            return users
        except sqlite3.Error as e:
            logger.error(f"Database error in get_all_users: {e}")
            return []

    async def update_main_menu(self, chat_id, context):
        status_text = "🟢 Active" if bot_active else "🔴 Inactive"
        keyboard = [
            [f"📝 Mark Attendance ({status_text})"],
            ["📋 My Attendance"],
            ["👤 My Profile"],
            ["📊 Report" if chat_id == ADMIN_ID else "ℹ️ Help"]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await context.bot.send_message(chat_id=chat_id, text="📋 Main Menu", reply_markup=reply_markup)

    # --- Core Features ---
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await self.update_main_menu(update.effective_chat.id, context)

    async def mark_attendance_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not bot_active:
            await update.message.reply_text("⚠️ Attendance bot is currently deactivated.")
            return ConversationHandler.END

        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if user_info:
            return await self.mark_attendance_immediate(update, context, user_info)
        else:
            if user_id in self.awaiting_registration:
                await update.message.reply_text("⏳ Please complete your registration first.")
                return
            self.awaiting_registration.add(user_id)
            await update.message.reply_text("📝 First time user? Let's register you!\n\nPlease enter your **Full Name**:")
            return NAME

    async def get_name(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['full_name'] = update.message.text.strip()
        await update.message.reply_text("✅ Name recorded!\n\nNow please enter your **Student ID**:")
        return STUDENT_ID

    async def get_student_id(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        student_id = update.message.text.strip()
        full_name = context.user_data['full_name']

        if not full_name or not student_id:
            await update.message.reply_text("❌ Both name and student ID are required.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END

        if self.student_id_exists(student_id):
            await update.message.reply_text("❌ This Student ID is already registered.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO users (user_id, full_name, student_id) VALUES (?, ?, ?)", (user_id, full_name, student_id))
            conn.commit()
            conn.close()

            self.awaiting_registration.discard(user_id)
            await update.message.reply_text(f"✅ Registration successful!\n\n**Name:** {full_name}\n**Student ID:** {student_id}\n\nYou can now mark your attendance.")
            await self.mark_attendance_immediate(update, context, (user_id, full_name, student_id))

        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Registration failed. Please try again.")
            self.awaiting_registration.discard(user_id)

        return ConversationHandler.END

    async def mark_attendance_immediate(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_info: tuple):
        if not bot_active:
            await update.message.reply_text("⚠️ Attendance bot is currently deactivated.")
            return

        user_id, full_name, student_id = user_info[0], user_info[1], user_info[2]
        today = datetime.date.today().isoformat()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if self.already_marked_today(user_id, today):
            await update.message.reply_text("✅ You've already marked attendance today!")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO attendance (user_id, full_name, student_id, date, time) VALUES (?, ?, ?, ?, ?)", (user_id, full_name, student_id, today, now))
            conn.commit()
            conn.close()

            self.save_to_excel(today, now, full_name, student_id)
            await update.message.reply_text(f"✅ Attendance marked successfully!\n\n**Time:** {now}\n**Name:** {full_name}\n**Student ID:** {student_id}")
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Failed to mark attendance. Please try again.")

    async def my_attendance(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if not user_info:
            await update.message.reply_text("❌ Please register first.")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT date, time FROM attendance WHERE user_id = ? ORDER BY date DESC, time DESC LIMIT 30", (user_id,))
            records = cursor.fetchall()
            conn.close()

            if not records:
                await update.message.reply_text("📭 No attendance records found.")
                return

            total_days = len(set(record[0] for record in records))
            msg = f"📋 **Your Attendance Records**\n\n**Name:** {user_info[1]}\n**Student ID:** {user_info[2]}\n**Total Days:** {total_days}\n\n"
            for date, time in records[:10]:
                msg += f"✅ {date} at {time}\n"
            if len(records) > 10:
                msg += f"\n... and {len(records) - 10} more records"

            await update.message.reply_text(msg, parse_mode="Markdown")
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error retrieving attendance records.")

    async def my_profile(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if not user_info:
            await update.message.reply_text("❌ Please register first.")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(DISTINCT date) FROM attendance WHERE user_id = ?", (user_id,))
            total_days = cursor.fetchone()[0]
            conn.close()

            msg = f"👤 **Your Profile**\n\n**Name:** {user_info[1]}\n**Student ID:** {user_info[2]}\n**Registered:** {user_info[3][:10]}\n**Total Attendance Days:** {total_days}"
            await update.message.reply_text(msg, parse_mode="Markdown")
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error retrieving profile.")

    async def report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Admin access required.")
            return

        today = datetime.date.today().isoformat()
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT full_name, student_id, time FROM attendance WHERE date = ? ORDER BY time", (today,))
            records = cursor.fetchall()
            conn.close()

            if not records:
                await update.message.reply_text("📭 No attendance records for today.")
                return

            msg = f"📊 **Today's Attendance Report**\n**Date:** {today}\n**Total Students:** {len(records)}\n\n"
            for i, (name, student_id, time) in enumerate(records, 1):
                msg += f"{i}. {name} ({student_id}) - {time}\n"

            self.generate_detailed_report(today, records)
            await update.message.reply_text(msg, parse_mode="Markdown")
            excel_file = f"attendance_logs/attendance_{today}.xlsx"
            if os.path.exists(excel_file):
                with open(excel_file, 'rb') as f:
                    await update.message.reply_document(document=f, filename=f"attendance_report_{today}.xlsx")
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error generating report.")

    # --- Delete Student ---
    async def delete_student_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can delete students.")
            return ConversationHandler.END
        
        await update.message.reply_text("🗑️ Please enter the Student ID you want to delete:")
        return DELETE_STUDENT_ID

    async def confirm_delete_student(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        student_id = update.message.text.strip()
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT full_name FROM users WHERE student_id = ?", (student_id,))
            user = cursor.fetchone()
            if not user:
                await update.message.reply_text(f"❌ No student found with ID {student_id}.")
                conn.close()
                return ConversationHandler.END

            cursor.execute("DELETE FROM attendance WHERE student_id = ?", (student_id,))
            cursor.execute("DELETE FROM users WHERE student_id = ?", (student_id,))
            conn.commit()
            conn.close()

            await update.message.reply_text(f"✅ Student '{user[0]}' (ID: {student_id}) deleted successfully!")
        except sqlite3.Error as e:
            logger.error(f"Database error in confirm_delete_student: {e}")
            await update.message.reply_text("⚠️ Database error occurred while deleting student.")
        return ConversationHandler.END

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        self.awaiting_registration.discard(update.effective_user.id)
        await update.message.reply_text("❌ Operation cancelled.")
        return ConversationHandler.END

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        help_text = (
            "<b>🤖 Attendance Bot Help</b>\n\n"
            "<b>For Students:</b>\n"
            "• <b>Mark Attendance</b> - Register and mark your daily attendance\n"
            "• <b>My Attendance</b> - View your attendance history\n"
            "• <b>My Profile</b> - See your profile and stats\n\n"
            "<b>For Admin:</b>\n"
            "• <b>Report</b> - View today's attendance report\n"
            "• <b>Delete Student</b> - Remove a student record\n\n"
            "<b>Features:</b>\n"
            "✅ Prevents duplicate registration\n"
            "✅ Prevents multiple markings per day\n"
            "✅ Secure database storage\n"
            "✅ Excel backups\n"
            "✅ User profiles with stats\n\n"
            "Need help? Contact admin: <a href='https://t.me/Neo_bx'>@Neo_bx</a>"
        )
        await update.message.reply_text(help_text, parse_mode="HTML", disable_web_page_preview=True)

    async def activate(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        global bot_active
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can activate the bot.")
            return
        bot_active = True
        await update.message.reply_text("✅ Attendance bot activated!")
        for user_id in self.get_all_users():
            try:
                await self.update_main_menu(user_id, context)
            except Exception as e:
                logger.error(f"Failed to update menu for user {user_id}: {e}")

    async def deactivate(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        global bot_active
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can deactivate the bot.")
            return
        bot_active = False
        await update.message.reply_text("⛔ Attendance bot deactivated!")
        for user_id in self.get_all_users():
            try:
                await self.update_main_menu(user_id, context)
            except Exception as e:
                logger.error(f"Failed to update menu for user {user_id}: {e}")

# --- Scheduler ---
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(60)

# --- Main ---
def main():
    init_database()
    bot = AttendanceBot()
    application = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & filters.Regex("📝 Mark Attendance"), bot.mark_attendance_start)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_name)],
            STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_student_id)],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)]
    )

    delete_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("delete_student", bot.delete_student_start)],
        states={DELETE_STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.confirm_delete_student)]},
        fallbacks=[CommandHandler("cancel", bot.cancel)]
    )

    application.add_handler(CommandHandler("start", bot.start))
    application.add_handler(CommandHandler("help", bot.help_command))
    application.add_handler(CommandHandler("activate", bot.activate))
    application.add_handler(CommandHandler("deactivate", bot.deactivate))
    application.add_handler(CommandHandler("report", bot.report))
    application.add_handler(conv_handler)
    application.add_handler(delete_conv_handler)

    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("📋 My Attendance"), bot.my_attendance))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("👤 My Profile"), bot.my_profile))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("📊 Report"), bot.report))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("ℹ️ Help"), bot.help_command))

    threading.Thread(target=run_scheduler, daemon=True).start()
    application.run_polling()

if __name__ == "__main__":
    main()
