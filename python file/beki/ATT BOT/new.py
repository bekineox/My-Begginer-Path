#!/usr/bin/env python3
"""
Attendance Telegram Bot (v20.8 compatible)
Save as: att_v4_fixed.py
Run: python att_v4_fixed.py
"""

import os
import datetime
import sqlite3
import logging
import threading
import time
from typing import Optional, List, Tuple

from openpyxl import Workbook, load_workbook
from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)
from telegram.constants import ParseMode

import schedule

# ---------------- Configuration ----------------
BOT_TOKEN = os.getenv("ATT_BOT_TOKEN") or "8469768440:AAFqN4d9cureq3g0f7N1-GGBM0CvZ69fK-U"
ADMIN_ID = int(os.getenv("ATT_ADMIN_ID") or 8191641097)  # change to your admin id
DB_PATH = "attendance.db"
LOGS_DIR = "attendance_logs"

# ---------------- Global bot status ----------------
bot_active = False

# ---------------- Conversation states ----------------
NAME, STUDENT_ID, DELETE_STUDENT_ID = range(3)

# ---------------- Logging ----------------
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


# ---------------- Database init ----------------
def init_database() -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            full_name TEXT NOT NULL,
            student_id TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            student_id TEXT,
            registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.commit()
    conn.close()


# ---------------- Helper functions ----------------
def now_iso() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class AttendanceBot:
    def __init__(self):
        self.awaiting_registration = set()

    def _connect(self):
        return sqlite3.connect(DB_PATH)

    def get_user_info(self, user_id: int) -> Optional[Tuple]:
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT user_id, full_name, student_id, registered_at FROM users WHERE user_id = ?", (user_id,))
            row = cur.fetchone()
            conn.close()
            return row
        except sqlite3.Error as e:
            logger.error("DB error (get_user_info): %s", e)
            return None

    def student_id_exists(self, student_id: str) -> bool:
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM users WHERE student_id = ?", (student_id,))
            exists = cur.fetchone() is not None
            conn.close()
            return exists
        except sqlite3.Error as e:
            logger.error("DB error (student_id_exists): %s", e)
            return False

    def already_marked_today(self, user_id: int, date: str) -> bool:
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM attendance WHERE user_id = ? AND date = ?", (user_id, date))
            marked = cur.fetchone() is not None
            conn.close()
            return marked
        except sqlite3.Error as e:
            logger.error("DB error (already_marked_today): %s", e)
            return False

    def save_to_excel(self, date: str, time_str: str, name: str, student_id: str) -> None:
        try:
            os.makedirs(LOGS_DIR, exist_ok=True)
            file_path = os.path.join(LOGS_DIR, f"attendance_{date}.xlsx")

            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws.append(["Date", "Time", "Full Name", "Student ID"])
            else:
                try:
                    wb = load_workbook(file_path)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Attendance"
                    ws.append(["Date", "Time", "Full Name", "Student ID"])

            ws.append([date, time_str, name, student_id])
            tmp_path = file_path + ".tmp"
            wb.save(tmp_path)
            wb.close()
            os.replace(tmp_path, file_path)
        except Exception as e:
            logger.error("Excel save error: %s", e)

    def generate_detailed_report(self, date: str, records: List[Tuple[str, str, str]]) -> str:
        """Creates a detailed report and returns the file path."""
        os.makedirs(LOGS_DIR, exist_ok=True)
        file_path = os.path.join(LOGS_DIR, f"attendance_{date}.xlsx")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.append(["No.", "Full Name", "Student ID", "Time", "Date"])
            for i, (name, student_id, time_str) in enumerate(records, start=1):
                ws.append([i, name, student_id, time_str, date])
            wb.save(file_path)
            wb.close()
        except Exception as e:
            logger.error("Excel report error: %s", e)
        return file_path

    def get_all_users(self) -> List[int]:
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT user_id FROM users")
            users = [r[0] for r in cur.fetchall()]
            conn.close()
            return users
        except sqlite3.Error as e:
            logger.error("DB error (get_all_users): %s", e)
            return []

    # ---------------- UI / commands ----------------
    async def update_main_menu(self, chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
        status_text = "🟢 Active" if bot_active else "🔴 Inactive"
        keyboard = [
            [f"📝 Mark Attendance ({status_text})"],
            ["📋 My Attendance"],
            ["👤 My Profile"],
            ["📊 Report" if chat_id == ADMIN_ID else "ℹ️ Help"],
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        try:
            # send or edit: here we send a short menu message
            await context.bot.send_message(chat_id=chat_id, text="📋 Main Menu", reply_markup=reply_markup)
        except Exception as e:
            logger.error("Failed to send main menu to %s: %s", chat_id, e)

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        await self.update_main_menu(update.effective_chat.id, context)

    async def mark_attendance_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not bot_active:
            await update.message.reply_text("⚠️ Attendance bot is currently deactivated.")
            return ConversationHandler.END

        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if user_info:
            # user_info is (user_id, full_name, student_id, registered_at)
            return await self.mark_attendance_immediate(update, context, user_info)
        else:
            if user_id in self.awaiting_registration:
                await update.message.reply_text("⏳ Please complete your registration first.")
                return ConversationHandler.END
            self.awaiting_registration.add(user_id)
            await update.message.reply_text("📝 First time user? Let's register you!\n\nPlease enter your Full Name:")
            return NAME

    async def get_name(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data["full_name"] = update.message.text.strip()
        await update.message.reply_text("✅ Name recorded!\n\nNow please enter your Student ID:")
        return STUDENT_ID

    async def get_student_id(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        student_id = update.message.text.strip()
        full_name = context.user_data.get("full_name", "").strip()

        if not full_name or not student_id:
            await update.message.reply_text("❌ Both name and student ID are required. Registration cancelled.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END

        if self.student_id_exists(student_id):
            await update.message.reply_text("❌ This Student ID is already registered. Registration cancelled.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END

        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("INSERT INTO users (user_id, full_name, student_id) VALUES (?, ?, ?)", (user_id, full_name, student_id))
            conn.commit()
            conn.close()

            self.awaiting_registration.discard(user_id)
            await update.message.reply_text(
                f"✅ Registration successful!\n\nName: {full_name}\nStudent ID: {student_id}\n\nYou can now mark your attendance."
            )
            # Immediately mark attendance
            return await self.mark_attendance_immediate(update, context, (user_id, full_name, student_id, now_iso()))
        except sqlite3.Error as e:
            logger.error("DB error (register): %s", e)
            await update.message.reply_text("❌ Registration failed. Please try again later.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END

    async def mark_attendance_immediate(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_info):
        if not bot_active:
            await update.message.reply_text("⚠️ Attendance bot is currently deactivated.")
            return ConversationHandler.END

        # user_info may be a tuple (user_id, full_name, student_id, registered_at) or (user_id, full_name, student_id)
        user_id = int(user_info[0])
        full_name = str(user_info[1])
        student_id = str(user_info[2])
        today = datetime.date.today().isoformat()
        time_str = now_iso()

        if self.already_marked_today(user_id, today):
            await update.message.reply_text("✅ You've already marked attendance today!")
            return ConversationHandler.END

        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO attendance (user_id, full_name, student_id, date, time) VALUES (?, ?, ?, ?, ?)",
                (user_id, full_name, student_id, today, time_str),
            )
            conn.commit()
            conn.close()

            # save to excel (append)
            self.save_to_excel(today, time_str, full_name, student_id)
            await update.message.reply_text(
                f"✅ Attendance marked successfully!\n\nTime: {time_str}\nName: {full_name}\nStudent ID: {student_id}",
                parse_mode=ParseMode.MARKDOWN,
            )
        except sqlite3.Error as e:
            logger.error("DB error (mark attendance): %s", e)
            await update.message.reply_text("❌ Failed to mark attendance. Please try again later.")
        return ConversationHandler.END

    async def my_attendance(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if not user_info:
            await update.message.reply_text("❌ Please register first (use Mark Attendance).")
            return

        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT date, time FROM attendance WHERE user_id = ? ORDER BY date DESC, time DESC LIMIT 100", (user_id,))
            records = cur.fetchall()
            conn.close()

            if not records:
                await update.message.reply_text("📭 No attendance records found.")
                return

            total_days = len(set(r[0] for r in records))
            msg_lines = [
                "📋 *Your Attendance Records*",
                f"*Name:* {user_info[1]}",
                f"*Student ID:* {user_info[2]}",
                f"*Total Days:* {total_days}",
                "",
            ]
            for date, time_str in records[:10]:
                msg_lines.append(f"✅ {date} at {time_str}")
            if len(records) > 10:
                msg_lines.append(f"\n... and {len(records) - 10} more records")

            await update.message.reply_text("\n".join(msg_lines), parse_mode=ParseMode.MARKDOWN)
        except sqlite3.Error as e:
            logger.error("DB error (my_attendance): %s", e)
            await update.message.reply_text("❌ Error retrieving attendance records.")

    async def my_profile(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        if not user_info:
            await update.message.reply_text("❌ Please register first (use Mark Attendance).")
            return

        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(DISTINCT date) FROM attendance WHERE user_id = ?", (user_id,))
            total_days = cur.fetchone()[0] or 0
            conn.close()

            registered_at = str(user_info[3])[:10] if user_info[3] else "N/A"
            msg = (
                f"👤 *Your Profile*\n\n"
                f"*Name:* {user_info[1]}\n"
                f"*Student ID:* {user_info[2]}\n"
                f"*Registered:* {registered_at}\n"
                f"*Total Attendance Days:* {total_days}"
            )
            await update.message.reply_text(msg, parse_mode=ParseMode.MARKDOWN)
        except sqlite3.Error as e:
            logger.error("DB error (my_profile): %s", e)
            await update.message.reply_text("❌ Error retrieving profile.")

    async def report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Admin access required.")
            return

        today = datetime.date.today().isoformat()
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT full_name, student_id, time FROM attendance WHERE date = ? ORDER BY time", (today,))
            records = cur.fetchall()
            conn.close()

            if not records:
                await update.message.reply_text("📭 No attendance records for today.")
                return

            msg_lines = [
                f"📊 *Today's Attendance Report*",
                f"*Date:* {today}",
                f"*Total Students:* {len(records)}",
                "",
            ]
            for i, (name, sid, time_str) in enumerate(records, start=1):
                msg_lines.append(f"{i}. {name} ({sid}) - {time_str}")

            # generate and send excel report
            file_path = self.generate_detailed_report(today, records)
            await update.message.reply_text("\n".join(msg_lines), parse_mode=ParseMode.MARKDOWN)
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    await update.message.reply_document(document=InputFile(f, filename=f"attendance_report_{today}.xlsx"))
        except sqlite3.Error as e:
            logger.error("DB error (report): %s", e)
            await update.message.reply_text("❌ Error generating report.")

    async def delete_student_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can delete students.")
            return ConversationHandler.END
        await update.message.reply_text("🗑️ Please enter the Student ID you want to delete:")
        return DELETE_STUDENT_ID

    async def confirm_delete_student(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        student_id = update.message.text.strip()
        try:
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT full_name FROM users WHERE student_id = ?", (student_id,))
            user = cur.fetchone()
            if not user:
                await update.message.reply_text(f"❌ No student found with ID {student_id}.")
                conn.close()
                return ConversationHandler.END

            cur.execute("DELETE FROM attendance WHERE student_id = ?", (student_id,))
            cur.execute("DELETE FROM users WHERE student_id = ?", (student_id,))
            conn.commit()
            conn.close()

            await update.message.reply_text(f"✅ Student '{user[0]}' (ID: {student_id}) deleted successfully!")
        except sqlite3.Error as e:
            logger.error("DB error (delete_student): %s", e)
            await update.message.reply_text("⚠️ Database error occurred while deleting student.")
        return ConversationHandler.END

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        self.awaiting_registration.discard(update.effective_user.id)
        await update.message.reply_text("❌ Operation cancelled.")
        return ConversationHandler.END

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        help_text = (
            "<b>🤖 Attendance Bot Help</b>\n\n"
            "<b>For Students:</b>\n"
            "• Mark Attendance - Register & mark your attendance\n"
            "• My Attendance - View your attendance history\n"
            "• My Profile - See your profile & stats\n\n"
            "Need help? Contact admin: <a href='https://t.me/Neo_bx'>@Neo_bx</a>"
        )
        await update.message.reply_text(help_text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)

    async def activate(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        global bot_active
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can activate the bot.")
            return
        bot_active = True
        await update.message.reply_text("✅ Attendance bot activated!")
        for user_id in self.get_all_users():
            try:
                await self.update_main_menu(user_id, context)
            except Exception as e:
                logger.error("Failed to update menu for user %s: %s", user_id, e)

    async def deactivate(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        global bot_active
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Only admin can deactivate the bot.")
            return
        bot_active = False
        await update.message.reply_text("⛔ Attendance bot deactivated!")
        for user_id in self.get_all_users():
            try:
                await self.update_main_menu(user_id, context)
            except Exception as e:
                logger.error("Failed to update menu for user %s: %s", user_id, e)


# ---------------- Scheduler ----------------
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(60)


# ---------------- Main ----------------
def main():
    init_database()
    bot = AttendanceBot()

    application = ApplicationBuilder().token(BOT_TOKEN).build()

    # Conversation for registration & marking
    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & filters.Regex(r"📝\s*Mark Attendance"), bot.mark_attendance_start)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_name)],
            STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_student_id)],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)],
        allow_reentry=True,
    )

    # Conversation for delete student (admin)
    delete_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("delete_student", bot.delete_student_start)],
        states={DELETE_STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.confirm_delete_student)]},
        fallbacks=[CommandHandler("cancel", bot.cancel)],
    )

    # Register handlers
    application.add_handler(CommandHandler("start", bot.start))
    application.add_handler(CommandHandler("help", bot.help_command))
    application.add_handler(CommandHandler("activate", bot.activate))
    application.add_handler(CommandHandler("deactivate", bot.deactivate))
    application.add_handler(CommandHandler("report", bot.report))
    application.add_handler(conv_handler)
    application.add_handler(delete_conv_handler)

    # Quick text handlers (from keyboard)
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"📋\s*My Attendance"), bot.my_attendance))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"👤\s*My Profile"), bot.my_profile))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"📊\s*Report"), bot.report))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"ℹ️\s*Help"), bot.help_command))

    # Start schedule thread (if you add scheduled jobs later)
    threading.Thread(target=run_scheduler, daemon=True).start()

    logger.info("Starting bot...")
    application.run_polling()


if __name__ == "__main__":
    main()
