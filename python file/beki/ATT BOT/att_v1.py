import os
import datetime
import sqlite3
import logging
import threading
import time
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)

import schedule

# === Configuration ===
BOT_TOKEN = "8469768440:AAFqN4d9cureq3g0f7N1-GGBM0CvZ69fK-U"
ADMIN_ID = 8191641097

# === Database Setup ===
DB_PATH = "attendance.db"

def init_database():
    """Initialize SQLite database"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            full_name TEXT NOT NULL,
            student_id TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            student_id TEXT,
            registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# === States for Conversation ===
NAME, STUDENT_ID = range(2)

# === Logging Setup ===
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

class AttendanceBot:
    def __init__(self):
        self.awaiting_registration = set()
    
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Send welcome message with keyboard"""
        keyboard = [
            ["📝 Mark Attendance"],
            ["📋 My Attendance"],
            ["👤 My Profile"],
            ["📊 Report" if update.effective_user.id == ADMIN_ID else "ℹ️ Help"]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

        user_info = self.get_user_info(update.effective_user.id)
        if user_info:
            welcome_msg = f"👋 Welcome back, {user_info[1]}!\nYour Student ID: {user_info[2]}"
        else:
            welcome_msg = (
    "👋 Welcome to the Section B Attendance Bot!\n"
    "This bot tracks attendance for Software Section B students.\n"
    "Please register first by selecting '📝 Mark Attendance' to start."
)


        await update.message.reply_text(
            f"{welcome_msg}\n\nChoose an option below:",
            reply_markup=reply_markup
        )

    def get_user_info(self, user_id: int) -> Optional[tuple]:
        """Get user info from database"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
            user = cursor.fetchone()
            conn.close()
            return user
        except sqlite3.Error as e:
            logger.error(f"Database error in get_user_info: {e}")
            return None


    async def mark_attendance_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start attendance marking process"""
        user_id = update.effective_user.id
        
        # Check if user is already registered
        user_info = self.get_user_info(user_id)
        if user_info:
            # User is registered, mark attendance directly
            return await self.mark_attendance_immediate(update, context, user_info)
        else:
            # User needs to register first
            if user_id in self.awaiting_registration:
                await update.message.reply_text("⏳ Please complete your registration first.")
                return
            
            self.awaiting_registration.add(user_id)
            await update.message.reply_text(
                "📝 First time user? Let's register you!\n\n"
                "Please enter your **Full Name**:"
            )
            return NAME

    async def get_name(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Store name and ask for student ID"""
        context.user_data['full_name'] = update.message.text.strip()
        await update.message.reply_text(
            "✅ Name recorded!\n\n"
            "Now please enter your **Student ID**:"
        )
        return STUDENT_ID

    async def get_student_id(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Store student ID and complete registration"""
        user_id = update.effective_user.id
        student_id = update.message.text.strip()
        full_name = context.user_data['full_name']
        
        # Validate input
        if not full_name or not student_id:
            await update.message.reply_text("❌ Both name and student ID are required.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END
        
        # Check if student ID already exists
        if self.student_id_exists(student_id):
            await update.message.reply_text("❌ This Student ID is already registered.")
            self.awaiting_registration.discard(user_id)
            return ConversationHandler.END
        
        # Save to database
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO users (user_id, full_name, student_id) VALUES (?, ?, ?)",
                (user_id, full_name, student_id)
            )
            conn.commit()
            conn.close()
            
            self.awaiting_registration.discard(user_id)
            await update.message.reply_text(
                f"✅ Registration successful!\n\n"
                f"**Name:** {full_name}\n"
                f"**Student ID:** {student_id}\n\n"
                "You can now mark your attendance."
            )
            
            # Mark attendance for today
            await self.mark_attendance_immediate(update, context, (user_id, full_name, student_id))
            
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Registration failed. Please try again.")
            self.awaiting_registration.discard(user_id)
        
        return ConversationHandler.END

    def student_id_exists(self, student_id: str) -> bool:
        """Check if student ID already exists"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM users WHERE student_id = ?", (student_id,))
            exists = cursor.fetchone() is not None
            conn.close()
            return exists
        except sqlite3.Error as e:
            logger.error(f"Database error in student_id_exists: {e}")
            return False

    async def mark_attendance_immediate(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_info: tuple):
        """Mark attendance for already registered user"""
        user_id, full_name, student_id = user_info[0], user_info[1], user_info[2]
        today = datetime.date.today().isoformat()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Check if already marked today
        if self.already_marked_today(user_id, today):
            await update.message.reply_text("✅ You've already marked attendance today!")
            return
        
        # Save to database
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO attendance (user_id, full_name, student_id, date, time) VALUES (?, ?, ?, ?, ?)",
                (user_id, full_name, student_id, today, now)
            )
            conn.commit()
            conn.close()
            
            # Also save to Excel for backup
            self.save_to_excel(today, now, full_name, student_id)
            
            await update.message.reply_text(
                f"✅ Attendance marked successfully!\n\n"
                f"**Time:** {now}\n"
                f"**Name:** {full_name}\n"
                f"**Student ID:** {student_id}"
            )
            
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Failed to mark attendance. Please try again.")

    def already_marked_today(self, user_id: int, date: str) -> bool:
        """Check if user already marked attendance today"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT 1 FROM attendance WHERE user_id = ? AND date = ?",
                (user_id, date)
            )
            marked = cursor.fetchone() is not None
            conn.close()
            return marked
        except sqlite3.Error as e:
            logger.error(f"Database error in already_marked_today: {e}")
            return False

    def save_to_excel(self, date: str, time: str, name: str, student_id: str):
        """Save attendance record to Excel file (safe & error-free)"""
        try:
            os.makedirs("attendance_logs", exist_ok=True)
            excel_file = f"attendance_logs/attendance_{date}.xlsx"

            if not os.path.isfile(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws.append(["Date", "Time", "Full Name", "Student ID"])
            else:
                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Attendance"
                    ws.append(["Date", "Time", "Full Name", "Student ID"])

            ws.append([date, time, name, student_id])
            temp_file = f"{excel_file}.tmp"
            wb.save(temp_file)
            wb.close()
            os.replace(temp_file, excel_file)

        except Exception as e:
            logger.error(f"Excel save error: {e}")

    async def my_attendance(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show user's attendance records"""
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        
        if not user_info:
            await update.message.reply_text("❌ Please register first using 'Mark Attendance'")
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT date, time FROM attendance WHERE user_id = ? ORDER BY date DESC, time DESC LIMIT 30",
                (user_id,)
            )
            records = cursor.fetchall()
            conn.close()
            
            if not records:
                await update.message.reply_text("📭 No attendance records found.")
                return
            
            total_days = len(set(record[0] for record in records))  # Unique dates
            
            msg = f"📋 **Your Attendance Records**\n\n"
            msg += f"**Name:** {user_info[1]}\n"
            msg += f"**Student ID:** {user_info[2]}\n"
            msg += f"**Total Days:** {total_days}\n\n"
            msg += "**Recent records:**\n"
            
            for date, time in records[:10]:  # Show last 10 records
                msg += f"✅ {date} at {time}\n"
            
            if len(records) > 10:
                msg += f"\n... and {len(records) - 10} more records"
            
            await update.message.reply_text(msg, parse_mode="Markdown")
            
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error retrieving attendance records.")

    async def my_profile(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show user profile"""
        user_id = update.effective_user.id
        user_info = self.get_user_info(user_id)
        
        if not user_info:
            await update.message.reply_text("❌ Please register first using 'Mark Attendance'")
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT COUNT(DISTINCT date) FROM attendance WHERE user_id = ?",
                (user_id,)
            )
            total_days = cursor.fetchone()[0]
            conn.close()
            
            msg = f"👤 **Your Profile**\n\n"
            msg += f"**Name:** {user_info[1]}\n"
            msg += f"**Student ID:** {user_info[2]}\n"
            msg += f"**Registered:** {user_info[3][:10]}\n"
            msg += f"**Total Attendance Days:** {total_days}"
            
            await update.message.reply_text(msg, parse_mode="Markdown")
            
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error retrieving profile.")

    async def report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Admin report - today's attendance"""
        user_id = update.effective_user.id
        if user_id != ADMIN_ID:
            await update.message.reply_text("⛔ Admin access required.")
            return
        
        today = datetime.date.today().isoformat()
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT full_name, student_id, time FROM attendance WHERE date = ? ORDER BY time",
                (today,)
            )
            records = cursor.fetchall()
            conn.close()
            
            if not records:
                await update.message.reply_text("📭 No attendance records for today.")
                return
            
            msg = f"📊 **Today's Attendance Report**\n"
            msg += f"**Date:** {today}\n"
            msg += f"**Total Students:** {len(records)}\n\n"
            
            for i, (name, student_id, time) in enumerate(records, 1):
                msg += f"{i}. {name} ({student_id}) - {time}\n"
            
            # Also generate Excel file
            self.generate_detailed_report(today, records)
            
            await update.message.reply_text(msg, parse_mode="Markdown")
            
            # Send Excel file
            excel_file = f"attendance_logs/attendance_{today}.xlsx"
            if os.path.exists(excel_file):
                with open(excel_file, 'rb') as f:
                    await update.message.reply_document(
                        document=f,
                        filename=f"attendance_report_{today}.xlsx"
                    )
                    
        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
            await update.message.reply_text("❌ Error generating report.")

    def generate_detailed_report(self, date: str, records: List[tuple]):
        """Generate detailed Excel report"""
        try:
            os.makedirs("attendance_logs", exist_ok=True)
            excel_file = f"attendance_logs/attendance_{date}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Headers
            ws.append(["No.", "Full Name", "Student ID", "Time", "Date"])
            
            # Data
            for i, (name, student_id, time) in enumerate(records, 1):
                ws.append([i, name, student_id, time, date])
            
            wb.save(excel_file)
        except Exception as e:
            logger.error(f"Excel report error: {e}")

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancel the current operation"""
        user_id = update.effective_user.id
        self.awaiting_registration.discard(user_id)
        await update.message.reply_text("❌ Operation cancelled.")
        return ConversationHandler.END

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show help message with safe HTML formatting"""
        help_text = (
            "<b>🤖 Attendance Bot Help</b>\n\n"
            "<b>For Students:</b>\n"
            "• <b>Mark Attendance</b> - Register and mark your daily attendance\n"
            "• <b>My Attendance</b> - View your attendance history\n"
            "• <b>My Profile</b> - See your profile and stats\n\n"
            "<b>For Admin:</b>\n"
            "• <b>Report</b> - View today's attendance report\n\n"
            "<b>Features:</b>\n"
            "✅ Prevents duplicate registration\n"
            "✅ Prevents multiple markings per day\n"
            "✅ Secure database storage\n"
            "✅ Excel backup files\n"
            "✅ User profiles with statistics\n\n"
            "Need help? Contact your administrator: "
            "<a href='https://t.me/Neo_bx'>@Neo_bx</a>"
        )

        await update.message.reply_text(help_text, parse_mode="HTML", disable_web_page_preview=True)





# === Background scheduler ===
def run_scheduler():
    """Run scheduled tasks"""
    while True:
        schedule.run_pending()
        time.sleep(60)

def main():
    """Start the bot"""
    # Initialize database
    init_database()
    
    # Create bot instance
    bot = AttendanceBot()
    
    # Build application
    application = ApplicationBuilder().token(BOT_TOKEN).build()
    
    # Conversation handler for registration
    conv_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex("📝 Mark Attendance"), bot.mark_attendance_start),
            CommandHandler("mark", bot.mark_attendance_start)
        ],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_name)],
            STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.get_student_id)],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)]
    )
    
    # Add handlers
    application.add_handler(CommandHandler("start", bot.start))
    application.add_handler(CommandHandler("help", bot.help_command))
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler("myattendance", bot.my_attendance))
    application.add_handler(CommandHandler("profile", bot.my_profile))
    application.add_handler(CommandHandler("report", bot.report))
    
    # Button handlers
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("📋 My Attendance"), bot.my_attendance))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("👤 My Profile"), bot.my_profile))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("📊 Report"), bot.report))
    application.add_handler(MessageHandler(filters.TEXT & filters.Regex("ℹ️ Help"), bot.help_command))
    
    # Start scheduler
    threading.Thread(target=run_scheduler, daemon=True).start()
    
    print("✅ Improved Attendance Bot is running...")
    application.run_polling()

if __name__ == '__main__':
    main()